"""
Excel 导出器
"""

from typing import Any
from .base import BaseExporter, ExportResult, ExportConfig


class ExcelExporter(BaseExporter):
    """Excel 格式导出"""
    
    def export(
        self,
        data: dict[str, Any],
        aimd_content: str = None,
        output_path: str = None,
    ) -> ExportResult:
        try:
            # 延迟导入，避免强制依赖
            try:
                import openpyxl
                from openpyxl.styles import Font, Alignment, PatternFill
            except ImportError:
                return ExportResult(
                    success=False, 
                    error="需要安装 openpyxl: pip install openpyxl"
                )
            
            wb = openpyxl.Workbook()
            
            # Sheet 1: 基本信息
            ws_info = wb.active
            ws_info.title = "基本信息"
            
            # 标题样式
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            
            ws_info.append(["字段", "值"])
            ws_info["A1"].font = header_font
            ws_info["B1"].font = header_font
            ws_info["A1"].fill = header_fill
            ws_info["B1"].fill = header_fill
            
            filtered = self._filter_data(data)
            row = 2
            for key, value in filtered.items():
                if not isinstance(value, (list, dict)):
                    ws_info.append([key, str(value) if value is not None else ""])
                    row += 1
            
            # 调整列宽
            ws_info.column_dimensions["A"].width = 30
            ws_info.column_dimensions["B"].width = 50
            
            # Sheet 2+: 表格数据
            for key, value in filtered.items():
                if isinstance(value, list) and len(value) > 0 and isinstance(value[0], dict):
                    ws_table = wb.create_sheet(title=key[:31])  # Excel sheet name max 31 chars
                    
                    # 表头
                    headers = list(value[0].keys())
                    ws_table.append(headers)
                    for col, _ in enumerate(headers, 1):
                        ws_table.cell(row=1, column=col).font = header_font
                        ws_table.cell(row=1, column=col).fill = header_fill
                    
                    # 数据行
                    for item in value:
                        ws_table.append([item.get(h, "") for h in headers])
            
            # 保存
            import io
            buffer = io.BytesIO()
            wb.save(buffer)
            content = buffer.getvalue()
            
            return self._save_or_return(content, output_path)
            
        except Exception as e:
            return ExportResult(success=False, error=str(e))


def export_excel(
    data: dict[str, Any],
    output_path: str = None,
    config: ExportConfig = None,
) -> ExportResult:
    """导出为 Excel 格式"""
    exporter = ExcelExporter(config)
    return exporter.export(data, output_path=output_path)
